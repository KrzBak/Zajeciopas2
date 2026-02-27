from qgis.core import *
from qgis.PyQt.QtGui import *
from qgis.PyQt.QtWidgets import *
import processing
from qgis.core import QgsProject, edit, QgsField
from PyQt5.QtCore import QVariant
from openpyxl import Workbook
from . import resources
from .form import message


class TestPlugin:

    def __init__(self, iface):
        self.iface = iface
        self.msg = None

    def initGui(self):
        self.action = QAction(
            QIcon(":/plugins/custom/icon.png"),
            "Zajęciopas",
            self.iface.mainWindow()
        )
        self.action.triggered.connect(self.run)

        self.iface.addToolBarIcon(self.action)
        self.iface.addPluginToMenu("&Home made", self.action)

    def unload(self):
        self.iface.removePluginMenu("&Home made", self.action)
        self.iface.removeToolBarIcon(self.action)

    def run(self):
        if self.msg is None:
            self.msg = message()

        try:
            self.msg.btnUruchom.clicked.disconnect()
            self.msg.btnExcel.clicked.disconnect()
            self.msg.cmbTabela.currentIndexChanged.disconnect()
        except:
            pass

        self.msg.btnUruchom.clicked.connect(self.uruchom_przycinanie)
        self.msg.btnExcel.clicked.connect(self.eksport_do_excel)
        
        self.msg.cmbTabela.currentIndexChanged.connect(self.przelacz_tabele)
    
        self.msg.show()

    # Przełączanie tabel

    def przelacz_tabele(self):

        if self.msg.cmbTabela.currentIndex() == 0:
            self.msg.tblWyniki.show()
            self.msg.tblKanalizacja.hide()
        else:
            self.msg.tblWyniki.hide()
            self.msg.tblKanalizacja.show()

    # Szerokość projektowanej kanalizacji

    def szerokosc_proj(self, ozn):

        ozn = str(ozn).upper()

        if "HDPE 50" in ozn:
            return 0.05
        elif "RHDPE 110" in ozn:
            return 0.11
        elif "HDPE 40" in ozn:
            return 0.04
        else:
            return 0.05
    
    def pobierz_model_glowny(self, model_full):

        if "DAC 2J" in model_full:
            return "DAC 2J"
        elif "DAC 12J" in model_full:
            return "DAC 2J"
        elif "DAC 4J" in model_full:
            return "DAC 2J"
        elif "ADSS" in model_full:
            return "ADSS"
        elif "MI-MKA" in model_full:
            return "MI-MKA"
        elif "MI-MKF" in model_full:
            return "MI-MKF"
        else:
            return "INNY"
    
    def pobierz_szerokosc(self, model_glowny):

        if model_glowny in ["ADSS", "DAC 2J", "DAC 12J", "DAC 4J"]:
            return 0.005
        elif model_glowny in ["MI-MKA", "MI-MKF"]:
            return 0.025
        else:
            return 0.005

    def uruchom_przycinanie(self):

        dzialki = self.msg.cmbDzialki.currentLayer()
        kable = self.msg.cmbKable.currentLayer()
        kan_istn = self.msg.cmbKanalizacja.currentLayer()
        kan_proj = self.msg.cmbKanalizacjaProj.currentLayer()

        if not dzialki or not kable:
            self.iface.messageBar().pushWarning(
                "Uwaga",
                "Wybierz warstwę działek i kabli."
            )
            return

        self.przytnij_kable(dzialki, kable, kan_istn, kan_proj)

    # Główna funkcja

    def przytnij_kable(self, warstwa_dzialki, warstwa_kable, warstwa_kan=None, warstwa_kan_proj=None):

        try:

            wynik = processing.run(
                "native:clip",
                {
                    'INPUT': warstwa_kable,
                    'OVERLAY': warstwa_dzialki,
                    'OUTPUT': 'memory:Kable_przyciete'
                }
            )

            layer = wynik['OUTPUT']
            
            # PRZYCINANIE KANALIZACJI PROJEKTOWANEJ DO DZIAŁEK

            kan_proj_clip = None

            if warstwa_kan_proj:

                wynik_kan = processing.run(
                    "native:clip",
                    {
                        'INPUT': warstwa_kan_proj,
                        'OVERLAY': warstwa_dzialki,
                        'OUTPUT': 'memory:Kanalizacja_proj_przycieta'
                    }
                )

                kan_proj_clip = wynik_kan['OUTPUT']
                QgsProject.instance().addMapLayer(kan_proj_clip)

            # BUFFER 0.5 m zabezpieczenia

            kan_buffer = None
            kan_proj_buffer = None

            if warstwa_kan:
                buf = processing.run(
                    "native:buffer",
                    {
                        'INPUT': warstwa_kan,
                        'DISTANCE': 0.1,
                        'DISSOLVE': True,
                        'OUTPUT': 'memory:buf1'
                    }
                )
                feats_buf = list(buf['OUTPUT'].getFeatures())
                if feats_buf:
                    kan_buffer = feats_buf[0].geometry()

            if warstwa_kan_proj:
                buf2 = processing.run(
                    "native:buffer",
                    {
                        'INPUT': warstwa_kan_proj,
                        'DISTANCE': 0.1,
                        'DISSOLVE': True,
                        'OUTPUT': 'memory:buf2'
                    }
                )
                feats_buf2 = list(buf2['OUTPUT'].getFeatures())
                if feats_buf2:
                    kan_proj_buffer = feats_buf2[0].geometry()

            with edit(layer):

                new_fields = []

                if layer.fields().indexOf("dl_full") == -1:
                    new_fields.append(QgsField("dl_full", QVariant.Double))

                if layer.fields().indexOf("pole_full") == -1:
                    new_fields.append(QgsField("pole_full", QVariant.Double))

                if new_fields:
                    layer.dataProvider().addAttributes(new_fields)
                    layer.updateFields()

                idx_full = layer.fields().indexOf("dl_full")
                idx_pole_full = layer.fields().indexOf("pole_full")

                for f in layer.getFeatures():

                    geom = f.geometry()
                    dl = geom.length()
                    dl_full = dl

                    typ = str(f["typ_elementu"]).lower() if "typ_elementu" in layer.fields().names() else ""

                    # Kabel napowietrzny → NIE ODEJMUJEMY kanalizacji

                    if typ == "kabel napowietrzny":

                        dl_full = dl

                    else:

                        if kan_buffer and geom.intersects(kan_buffer):
                            g = geom.difference(kan_buffer)
                            dl_full = g.length() if not g.isEmpty() else 0

                        if kan_proj_buffer and geom.intersects(kan_proj_buffer):
                            g = geom.difference(kan_proj_buffer)
                            dl_full = g.length() if not g.isEmpty() else dl_full

                    model_full = str(f['model_kabla'])
                    model_glowny = self.pobierz_model_glowny(model_full)
                    szer = self.pobierz_szerokosc(model_glowny)
                    
                    pole_full = dl_full * szer

                    layer.changeAttributeValue(f.id(), idx_full, dl_full)
                    layer.changeAttributeValue(f.id(), idx_pole_full, pole_full)

            QgsProject.instance().addMapLayer(layer)
            
            # WARSTWA TYMCZASOWA - KABLE BEZ KANALIZACJI

            bez_kan_layer = QgsVectorLayer(
                "LineString?crs=" + layer.crs().authid(),
                "Kable_przyciete_bez_kanalizacji",
                "memory"
            )

            prov = bez_kan_layer.dataProvider()
            prov.addAttributes(layer.fields())
            bez_kan_layer.updateFields()

            features_new = []

            for f in layer.getFeatures():

                geom = f.geometry()
                typ = str(f["typ_elementu"]).lower() if "typ_elementu" in layer.fields().names() else ""

                # Kabel napowietrzny bez kanalizacji kanalizacji

                if typ != "kabel napowietrzny":

                    if kan_buffer and geom.intersects(kan_buffer):
                        geom = geom.difference(kan_buffer)

                    if kan_proj_buffer and geom.intersects(kan_proj_buffer):
                        geom = geom.difference(kan_proj_buffer)

                if not geom.isEmpty():

                    new_feat = QgsFeature()
                    new_feat.setGeometry(geom)
                    new_feat.setAttributes(f.attributes())
                    features_new.append(new_feat)

            prov.addFeatures(features_new)
            bez_kan_layer.updateExtents()

            QgsProject.instance().addMapLayer(bez_kan_layer)

            # TABELA KABLI

            feats = list(layer.getFeatures())
            table = self.msg.tblWyniki

            headers = [
                "ID",
                "Model kabla",
                "Typ elementu",
                "Długość [m]",
                "Dlugosc Full",
                "Pole Full"
            ]

            table.setRowCount(len(feats))
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            
            idx_typ = layer.fields().indexOf("typ_elementu")

            for row, feat in enumerate(feats):

                model_full = str(feat['model_kabla'])
                model_glowny = self.pobierz_model_glowny(model_full)
                
                table.setItem(row, 0, QTableWidgetItem(str(feat.id())))
                table.setItem(row, 1, QTableWidgetItem(model_glowny))
                if idx_typ != -1:
                    table.setItem(row, 2, QTableWidgetItem(str(feat['typ_elementu'])))
                else:
                    table.setItem(row, 2, QTableWidgetItem("—"))
                table.setItem(row, 3, QTableWidgetItem(f"{feat.geometry().length():.2f}"))
                table.setItem(row, 4, QTableWidgetItem(f"{feat['dl_full']:.2f}"))
                table.setItem(row, 5, QTableWidgetItem(f"{feat['pole_full']:.4f}"))

            table.resizeColumnsToContents()

            # TABELA PROJEKTOWANEJ KANALIZACJI

            if warstwa_kan_proj:

                feats2 = list(kan_proj_clip.getFeatures())
                table2 = self.msg.tblKanalizacja

                headers2 = ["typ_elementu", "oznaczenie", "Długość [m]", "Pole [m2]"]

                table2.setRowCount(len(feats2))
                table2.setColumnCount(len(headers2))
                table2.setHorizontalHeaderLabels(headers2)

                for row, f in enumerate(feats2):

                    dl = f.geometry().length()
                    ozn = f["oznaczenie"] if "oznaczenie" in warstwa_kan_proj.fields().names() else ""
                    typ = f["typ_elementu"] if "typ_elementu" in warstwa_kan_proj.fields().names() else ""

                    szer = self.szerokosc_proj(ozn)
                    pole = dl * szer

                    table2.setItem(row, 0, QTableWidgetItem(str(typ)))
                    table2.setItem(row, 1, QTableWidgetItem(str(ozn)))
                    table2.setItem(row, 2, QTableWidgetItem(f"{dl:.2f}"))
                    table2.setItem(row, 3, QTableWidgetItem(f"{pole:.4f}"))

                table2.resizeColumnsToContents()

            self.iface.messageBar().pushSuccess(
                "OK",
                "Przycinanie zakończone."
            )

        except Exception as e:
            self.iface.messageBar().pushCritical(
                "Błąd",
                str(e)
            )



    # Export do excela
    
    
    def eksport_do_excel(self):

        path, _ = QFileDialog.getSaveFileName(
            self.msg,
            "Zapisz plik Excel",
            "",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wyniki"

        table = self.msg.tblWyniki if self.msg.cmbTabela.currentIndex() == 0 else self.msg.tblKanalizacja

        for col in range(table.columnCount()):
            header = table.horizontalHeaderItem(col).text()
            sheet.cell(row=1, column=col + 1, value=header)

        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    sheet.cell(row=row + 2, column=col + 1, value=item.text())

        workbook.save(path)

        self.iface.messageBar().pushSuccess(
            "Sukces",
            "Dane wyeksportowane do Excel."
        )
        
    